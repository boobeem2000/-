import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Create sheets
dashboard = wb.create_sheet('Dashboard', 0)
mission_log = wb.create_sheet('Mission Log', 1)
instructions = wb.create_sheet('Instructions', 2)

# Define colors and fonts
MINT_GREEN = 'E8F5E8'
SUNRISE_ORANGE = 'FFE4B5'
LIGHT_BLUE = 'E6F3FF'

title_font = Font(name='メイリオ', size=20, bold=True, color='2E8B57')
subtitle_font = Font(name='メイリオ', size=14, bold=True, color='FF8C00')
body_font = Font(name='メイリオ', size=11)

mint_fill = PatternFill(start_color=MINT_GREEN, end_color=MINT_GREEN, fill_type='solid')
orange_fill = PatternFill(start_color=SUNRISE_ORANGE, end_color=SUNRISE_ORANGE, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')

# DASHBOARD SHEET
dashboard.merge_cells('A1:J3')
title_cell = dashboard['A1']
title_cell.value = '🎮 Atsuko式スピーキング再起動ゲーム\n〜7 Days to Restart Your English Voice〜'
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = mint_fill

dashboard['A5'] = '📊 進捗状況'
dashboard['A5'].font = subtitle_font
dashboard['A5'].fill = orange_fill

dashboard['A7'] = '総得点:'
dashboard['B7'] = '=SUM("Mission Log"!F2:F8)'
dashboard['C7'] = '/ 21点'
dashboard['A7'].font = body_font
dashboard['B7'].font = Font(name='メイリオ', size=14, bold=True, color='2E8B57')
dashboard['C7'].font = body_font

dashboard['A9'] = '=IF(B7=21,"🏆 LEVEL UP! おめでとうございます！","頑張って続けましょう！")'
dashboard['A9'].font = Font(name='メイリオ', size=12, bold=True, color='FF8C00')
dashboard.merge_cells('A9:J9')

dashboard['A11'] = '🎯 Today Mission'
dashboard['A11'].font = subtitle_font
dashboard['A11'].fill = light_blue_fill
dashboard.merge_cells('A11:J11')

dashboard['A13'] = 'Mission Logシートで毎日のミッションに挑戦してください！'
dashboard['A13'].font = body_font
dashboard.merge_cells('A13:J13')

dashboard['A15'] = '📋 使い方'
dashboard['A15'].font = subtitle_font
dashboard['A15'].fill = orange_fill

instructions_text = [
    '1. 毎日1つのミッションに挑戦',
    '2. 完了したらMission Logに記録',
    '3. 得点を入力して進捗を確認',
    '4. 7日間完了でLEVEL UP！'
]

for i, instruction in enumerate(instructions_text):
    dashboard[f'A{17+i}'] = instruction
    dashboard[f'A{17+i}'].font = body_font

for col in range(1, 11):
    dashboard.column_dimensions[get_column_letter(col)].width = 15

# MISSION LOG SHEET
headers = ['Day', 'ミッションタイトル', '指令内容', '英語で話す例', '今日の気づき', '🌟得点', 'ごほうび']

for col, header in enumerate(headers, 1):
    cell = mission_log.cell(row=1, column=col)
    cell.value = header
    cell.font = subtitle_font
    cell.fill = mint_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

missions = [
    [1, 'Restartボタンを押せ！', '「今週から再開します」と言おう', 'I am restarting my English journey today.', '', '3', '☕ コーヒーご褒美'],
    [2, '3文ストーリー職人', '朝の出来事を3文で語れ', 'I overslept, missed the bus, but smiled anyway.', '', '4', '🌈 5分早上がり'],
    [3, '翻訳禁止チャレンジ', '日本語を使わずイメージで話す', 'My brain is still sleeping.', '', '3', '🍫 チョコ解禁'],
    [4, '録音サバイバー', '自分の声を聴いて褒める', '（自由トピック）', '', '4', '🎬 推し動画10分'],
    [5, '失敗ヒーロー', '英語で失敗談を語れ', 'I said I am exciting to my boss 😅', '', '4', '🧘‍♀️ 夜ヨガ'],
    [6, '上司に報告ミッション', '報告するつもりで話す', 'I checked the report. It is ready.', '', '3', '🛁 バスソルト'],
    [7, 'Final Quest✨', 'I am deserved to do itを語る', 'I feel ready to speak English anytime.', '', '5', '👏 自分に拍手']
]

for row, mission in enumerate(missions, 2):
    for col, value in enumerate(mission, 1):
        cell = mission_log.cell(row=row, column=col)
        cell.value = value
        cell.font = body_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if col == 2:
            cell.font = Font(name='メイリオ', size=11, bold=True)
        elif col == 6:
            cell.font = Font(name='メイリオ', size=11, bold=True, color='2E8B57')
        elif col == 7:
            cell.font = Font(name='メイリオ', size=10, color='FF8C00')

column_widths = [6, 20, 25, 30, 20, 8, 15]
for i, width in enumerate(column_widths, 1):
    mission_log.column_dimensions[get_column_letter(i)].width = width

# INSTRUCTIONS SHEET
instructions.merge_cells('A1:F2')
title_cell = instructions['A1']
title_cell.value = '📖 ゲームルール & 使い方'
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = mint_fill

instructions['A4'] = '🎯 ゲームコンセプト'
instructions['A4'].font = subtitle_font
instructions['A4'].fill = orange_fill

concept_text = [
    '「努力」ではなく「冒険」──学習をRPG化する',
    '英語への抵抗感を「遊び心」で溶かす',
    '進むたびに自信が湧く、温かくポップな世界観',
    '1日3分の小さな成功を、7回積み重ねる'
]

for i, text in enumerate(concept_text):
    instructions[f'A{6+i}'] = text
    instructions[f'A{6+i}'].font = body_font

instructions['A11'] = '🎮 遊び方'
instructions['A11'].font = subtitle_font
instructions['A11'].fill = light_blue_fill

how_to_play = [
    '1. 毎日1つのミッションに挑戦してください',
    '2. 英語で話す例を参考に、実際に声に出して練習',
    '3. 完了したら「今日の気づき」欄に感想を記入',
    '4. 得点を入力（1-5点で自己評価）',
    '5. ごほうびを自分に与えてください！',
    '6. 7日間完了でLEVEL UP達成！'
]

for i, step in enumerate(how_to_play):
    instructions[f'A{13+i}'] = step
    instructions[f'A{13+i}'].font = body_font

instructions['A20'] = '⭐ 得点システム'
instructions['A20'].font = subtitle_font
instructions['A20'].fill = orange_fill

scoring_text = [
    '1点: 挑戦した（偉い！）',
    '2点: 少しできた',
    '3点: まあまあできた',
    '4点: よくできた',
    '5点: 完璧！'
]

for i, score in enumerate(scoring_text):
    instructions[f'A{22+i}'] = score
    instructions[f'A{22+i}'].font = body_font

instructions['A28'] = '💡 成功のコツ'
instructions['A28'].font = subtitle_font
instructions['A28'].fill = light_blue_fill

tips = [
    '• 完璧を求めず、楽しむことを最優先に',
    '• 小さな成功を積み重ねることが大切',
    '• 失敗も学習の一部として受け入れる',
    '• 毎日続けることで習慣化を目指す',
    '• 自分を褒めることを忘れずに！'
]

for i, tip in enumerate(tips):
    instructions[f'A{30+i}'] = tip
    instructions[f'A{30+i}'].font = body_font

for col in range(1, 7):
    instructions.column_dimensions[get_column_letter(col)].width = 20

wb.save('Atsuko式スピーキング再起動ゲーム.xlsx')
print('Excel workbook created successfully!')

